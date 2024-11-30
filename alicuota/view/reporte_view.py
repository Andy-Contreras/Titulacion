import openpyxl
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from django.utils.timezone import datetime
from django.views.generic import TemplateView
import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from alicuota.models import CabFactura, DetFactura


class ReporteListView(LoginRequiredMixin, TemplateView):
    template_name = 'reporte.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Reportes de Pagos'

        # Obtener fechas del GET request
        fecha_inicio = self.request.GET.get('fecha_inicio', '')
        fecha_fin = self.request.GET.get('fecha_fin', '')

        # Inicializar valores para facturas y total
        context['facturas'] = []
        context['total'] = 0

        if fecha_inicio and fecha_fin:
            try:
                # Convertir las fechas a objetos de tipo fecha
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

                # Filtrar las cabeceras de factura (CabFactura) por el rango de fechas
                cabfacturas = CabFactura.objects.filter(
                    fecha_creacion__range=(fecha_inicio_obj, fecha_fin_obj)
                )

                # Calcular el total de pagos de las facturas
                total = cabfacturas.aggregate(total_pagos=Sum('monto')).get('total_pagos') or 0

                # Pasar las facturas y el total al contexto
                context['facturas'] = cabfacturas
                context['total'] = total
            except ValueError:
                context['error'] = "Formato de fecha inválido"

        # Siempre pasar las fechas seleccionadas al contexto
        context['fecha_inicio'] = fecha_inicio
        context['fecha_fin'] = fecha_fin

        return context

    def get(self, request, *args, **kwargs):
        # Verificar si se solicita la descarga del Excel
        if 'descargar_excel' in request.GET:
            return self.generar_excel()

        return super().get(request, *args, **kwargs)

    def generar_excel(self):
        # Obtener las fechas del GET request
        fecha_inicio = self.request.GET.get('fecha_inicio', '')
        fecha_fin = self.request.GET.get('fecha_fin', '')

        if fecha_inicio and fecha_fin:
            try:
                # Convertir las fechas a objetos de tipo fecha
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

                # Filtrar las cabeceras de factura
                cabfacturas = CabFactura.objects.filter(
                    fecha_creacion__range=(fecha_inicio_obj, fecha_fin_obj)
                )

                # Crear una lista de diccionarios con los datos
                data = [{
                    'Fecha': factura.fecha_creacion,
                    'Nombre': factura.id_residentes.nombre,
                    'Valor Pagado': factura.monto
                } for factura in cabfacturas]

                # Crear un DataFrame de pandas
                df = pd.DataFrame(data)

                # Agregar fila para el total
                total = cabfacturas.aggregate(total_pagos=Sum('monto')).get('total_pagos') or 0
                df.loc[len(df)] = ['Total', '', total]

                # Preparar la respuesta HTTP para el archivo Excel
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="reporte_{fecha_inicio}_{fecha_fin}.xlsx"'

                with pd.ExcelWriter(response, engine='openpyxl') as writer:
                    workbook = writer.book

                    # Crear una hoja nueva
                    worksheet = workbook.create_sheet('Reporte de Pagos')

                    # Añadir el título en la primera fila
                    worksheet.append([f"Reporte de Pagos desde {fecha_inicio} hasta {fecha_fin}"])
                    worksheet.append([])  # Añadir una fila vacía

                    # Añadir los encabezados de las columnas
                    for i, column in enumerate(df.columns, start=1):
                        worksheet.cell(row=3, column=i, value=column)

                    # Añadir los datos debajo de los encabezados
                    for row_idx, row_data in enumerate(df.values, start=4):
                        for col_idx, cell_value in enumerate(row_data, start=1):
                            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

                    # Estilo de los encabezados
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    header_border = Border(
                        top=Side(style='thin', color='000000'),
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )

                    for col in range(1, len(df.columns) + 1):
                        max_length = len(df.columns[col - 1])  # Inicializa con el ancho del encabezado
                        column_letter = openpyxl.utils.get_column_letter(col)

                        # Iterar sobre las filas para determinar el ancho máximo
                        for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, min_col=col, max_col=col):
                            for cell in row:
                                if cell.value:  # Verifica si el valor no es None
                                    max_length = max(max_length, len(str(cell.value)))

                        # Ajustar el ancho de la columna
                        adjusted_width = max_length + 2  # Agrega un pequeño margen
                        worksheet.column_dimensions[column_letter].width = adjusted_width

                    # Definir la tabla de Excel
                    tab = Table(displayName="Reporte", ref=f"A3:{openpyxl.utils.get_column_letter(len(df.columns))}{worksheet.max_row}")

                    # Agregar el estilo de la tabla
                    style = TableStyleInfo(
                        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                        showRowStripes=True, showColumnStripes=True
                    )
                    tab.tableStyleInfo = style

                    # Agregar la tabla a la hoja
                    worksheet.add_table(tab)

                return response
            except ValueError:
                return HttpResponse("Formato de fecha inválido", status=400)

        return HttpResponse("Fechas no proporcionadas", status=400)


