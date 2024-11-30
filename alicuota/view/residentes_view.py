
import openpyxl
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, DeleteView, UpdateView
from django.contrib import messages
from alicuota.forms import ResidenteForm
from alicuota.models import *
from django.urls import reverse_lazy
from django.shortcuts import redirect
from django.db.models import Q
from django.db.models import CharField
from django.views.generic import TemplateView
import pandas as pd
from django.http import HttpResponse
from django.db.models import F, Value
from django.db.models.functions import Concat
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

class ResidenteListView(LoginRequiredMixin, ListView):
    template_name = 'Residente/lista_residente.html'
    model = Residente
    queryset = Residente.objects.filter(activo=True)

    def get_queryset(self):
        query = self.request.GET.get("query")
        if query:
            return self.model.objects.filter(
                (Q(nombre__icontains=query) | Q(cedula__icontains=query)) & Q(activo=True)
            )
        return super().get_queryset()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Listado de Residente'
        context['btn_crear'] = 'Añadir'
        context['buscar'] = 'Ingrese el nomrbre'
        context['url_crear'] = '/residente_crear'
        context['btn_eliminar'] = '/residente_eliminar'
        context['btn_actualizar'] = '/residente_actualizar'
        context['query'] = self.request.GET.get("query")
        context['listar_url'] = '/residente_lista/'
        return context


class ResidenteCreateView(LoginRequiredMixin, CreateView):
    template_name = 'Residente/crear_residente.html'
    model = Residente
    form_class = ResidenteForm
    success_url = reverse_lazy('residente_lista')
    context_object_name = 'residentes'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'NUEVO REGISTRO'
        context['action_save'] = '/residente_crear/'
        context['listar_url'] = '/residente_lista/'
        context['cancel_url'] = reverse_lazy('residente_lista')
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        patient = self.object
        save_audit(self.request, patient, action='A')
        messages.success(self.request, f"Éxito al Crear el Residente {patient.nombre}.")
        print("mande mensaje")
        return response


class ResidenteDeleteView(LoginRequiredMixin, DeleteView):
    template_name = 'Residente/eliminar_residente.html'
    model = Residente
    success_url = reverse_lazy('residente_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'ELIMINAR RESIDENTE'
        context['action_save'] = self.request.path
        context['cancel_url'] = reverse_lazy('residente_lista')
        return context

    def delete(self, request, *args, **kwargs):
        # Aquí se maneja la eliminación en sí, llamando a la función base
        response = super().delete(request, *args, **kwargs)
        return response

    def post(self, request, pk, *args, **kwargs):
        # Obtiene el residente a eliminar
        residente = Residente.objects.get(id=pk)
        # Guardar auditoría
        save_audit(request, residente, action='E')  # Registra la acción de eliminación lógica
        # Eliminar lógicamente al residente
        residente.activo = False
        residente.save()
        # Desasociar al residente de la vivienda
        Vivienda.objects.filter(residente=residente).update(residente=None)
        # Mensaje de confirmación
        messages.success(request, f"Residente {residente.nombre} eliminado lógicamente y desasociado de la vivienda.")
        # Redirección a la lista de residentes
        return redirect('residente_lista')


class ResidenteUpdateView(LoginRequiredMixin, UpdateView):
    template_name = 'Residente/crear_residente.html'
    model = Residente
    form_class = ResidenteForm
    success_url = reverse_lazy('residente_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'ACTUALIZAR RESIDENTE'
        context['action_save'] = self.request.path
        context['cancel_url'] = reverse_lazy('residente_lista')
        return context

    def form_valid(self, form):
        # Guardar el formulario y realizar la auditoría
        response = super().form_valid(form)
        residente = self.object

        # Llamar al metodo de auditoría
        save_audit(self.request, residente, action='M')

        # Mostrar un mensaje de éxito al usuario
        messages.success(self.request, f"Éxito al modificar el residente: {residente.nombre}.")

        return response

class ReporteResidenteView(TemplateView):
    template_name = 'reporte_residente.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Reporte Residente'

        # Obtener fechas del GET request
        fecha_inicio = self.request.GET.get('fecha_inicio', '')
        fecha_fin = self.request.GET.get('fecha_fin', '')

        # Inicializar valores para facturas y total
        context['residente'] = []

        if fecha_inicio and fecha_fin:
            try:
                # Convertir las fechas a objetos de tipo fecha
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

                # Realizar la consulta de facturas con el ORM
                residente = Residente.objects.filter(
                    status=1,
                    vivienda__cabalicuota__saldo_pendiente__gt=0,
                    vivienda__cabalicuota__fecha_creacion__range=(fecha_inicio_obj, fecha_fin_obj)
                ).annotate(
                    manzana_villa=Concat(
                        F('vivienda__villa__manzana__codigo_manzana'),
                        Value(' - '),
                        F('vivienda__villa__numero_villa'),
                        output_field=CharField()
                    ),
                    periodo=F('vivienda__cabalicuota__periodo'),
                    porcentaje=F('vivienda__cabalicuota__Interes__porcentaje'),
                    saldo_financiar=F('vivienda__cabalicuota__saldo_financiar'),
                    saldo_pendiente=F('vivienda__cabalicuota__saldo_pendiente'),
                    fecha_creacion=F('vivienda__cabalicuota__fecha_creacion')
                ).values(
                    'id', 'nombre', 'manzana_villa', 'periodo', 'porcentaje', 'saldo_financiar', 'saldo_pendiente', 'fecha_creacion'
                )

                # Pasar las facturas y el total al contexto
                context['residente'] = residente
            except ValueError:
                context['error'] = "Formato de fecha inválido"

        # Siempre pasar las fechas seleccionadas al contexto
        context['fecha_inicio'] = fecha_inicio
        context['fecha_fin'] = fecha_fin

        return context

    def get(self, request, *args, **kwargs):
        # Verificar si se solicita descargar el Excel
        if 'descargar_excel' in request.GET:
            return self.generar_excel()

        return super().get(request, *args, **kwargs)

    def generar_excel(self):
        # Obtener fechas del GET request
        fecha_inicio = self.request.GET.get('fecha_inicio', '')
        fecha_fin = self.request.GET.get('fecha_fin', '')

        if fecha_inicio and fecha_fin:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

                # Consultar los datos
                residente = Residente.objects.filter(
                    status=1,
                    vivienda__cabalicuota__saldo_pendiente__gt=0,
                    vivienda__cabalicuota__fecha_creacion__range=(fecha_inicio_obj, fecha_fin_obj)
                ).annotate(
                    manzana_villa=Concat(
                        F('vivienda__villa__manzana__codigo_manzana'),
                        Value(' - '),
                        F('vivienda__villa__numero_villa'),
                        output_field=CharField()
                    ),
                    periodo=F('vivienda__cabalicuota__periodo'),
                    porcentaje=F('vivienda__cabalicuota__Interes__porcentaje'),
                    saldo_financiar=F('vivienda__cabalicuota__saldo_financiar'),
                    saldo_pendiente=F('vivienda__cabalicuota__saldo_pendiente'),
                    fecha_creacion=F('vivienda__cabalicuota__fecha_creacion')
                ).values(
                    'id', 'nombre', 'manzana_villa', 'periodo', 'porcentaje', 'saldo_financiar', 'saldo_pendiente', 'fecha_creacion'
                )

                # Convertir los datos en un DataFrame de pandas
                df = pd.DataFrame(list(residente))

                df.rename(columns={
                    'nombre': 'Nombre',
                    'manzana_villa': 'Manzana - Villa',
                    'periodo': 'Periodo',
                    'porcentaje': 'Porcentaje Interés',
                    'saldo_financiar': 'Saldo a Financiar',
                    'saldo_pendiente': 'Saldo Pendiente',
                    'fecha_creacion': 'Fecha Creación'
                }, inplace=True)

                if 'id' in df.columns:
                    df.drop(columns=['id'], inplace=True)

                # Preparar la respuesta HTTP para el archivo Excel
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response[
                    'Content-Disposition'] = f'attachment; filename="Reporte_Residente_{fecha_inicio}_{fecha_fin}.xlsx"'

                # Escribir el archivo Excel con un título
                with pd.ExcelWriter(response, engine='openpyxl') as writer:
                    workbook = writer.book

                    # Crear una hoja nueva con nombre específico
                    worksheet = workbook.create_sheet('Reporte Residente')

                    # Añadir título en la primera fila
                    worksheet.append([f"Reporte de Residentes desde {fecha_inicio} hasta {fecha_fin}"])
                    worksheet.append([])  # Fila vacía

                    # Añadir los encabezados de las columnas
                    for i, column in enumerate(df.columns, start=1):
                        worksheet.cell(row=3, column=i, value=column)

                    # Añadir los datos debajo de los encabezados
                    for row_idx, row_data in enumerate(df.values, start=4):
                        for col_idx, cell_value in enumerate(row_data, start=1):
                            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

                    # Estilo de los encabezados
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    header_border = Border(
                        top=Side(style='thin', color='000000'),
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )

                    for col in range(1, len(df.columns) + 1):
                        max_length = len(df.columns[col - 1])  # Inicializa con el ancho del encabezado
                        column_letter = openpyxl.utils.get_column_letter(col)

                        # Iterar sobre las filas para determinar el ancho máximo
                        for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, min_col=col, max_col=col):
                            for cell in row:
                                if cell.value:  # Verifica si el valor no es None
                                    max_length = max(max_length, len(str(cell.value)))

                        # Ajustar el ancho de la columna
                        adjusted_width = max_length + 2  # Agrega un pequeño margen
                        worksheet.column_dimensions[column_letter].width = adjusted_width

                    # Definir la tabla de Excel
                    tab = Table(displayName="Reporte",
                                ref=f"A3:{openpyxl.utils.get_column_letter(len(df.columns))}{worksheet.max_row}")

                    # Agregar el estilo de la tabla
                    style = TableStyleInfo(
                        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                        showRowStripes=True, showColumnStripes=True
                    )
                    tab.tableStyleInfo = style

                    # Agregar la tabla a la hoja
                    worksheet.add_table(tab)

                return response

            except ValueError:
                return HttpResponse("Error al generar el Excel. Formato de fecha inválido.", status=400)

        return HttpResponse("Debe proporcionar un rango de fechas válido.", status=400)

